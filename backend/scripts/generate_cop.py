#!/usr/bin/env python3
"""
Generador COP (Costo Operativo Pacientes) para CDIEM - Centro Oncologico
Genera archivo Excel (.xlsx) con 40 hojas:
  Hoja 1  — Resumen
  Hoja 2  — Descripcion
  Hoja 3  — Hora Sillon
  Hoja 4  — Prevision
  Hojas 5-40 — Hoja1 ... Hoja36  (datos se definiran posteriormente)

Uso: python3 generate_cop.py <input.json> <output.xlsx>
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Opciones de prevision (dropdown hoja Prevision) ───────────────────────────
PREVISIONES = [
    "Fonasa",
    "Isapre Colmena",
    "Isapre Banmedica",
    "Isapre Vida Tres",
    "Isapre Masvida",
    "Isapre Consalud",
    "Isapre Fundacion",
    "Isalud",
    "Isapre Cruz Blanca",
    "Cruz del Norte",
    "Isapre Esencial",
]

# ── Paleta de colores ─────────────────────────────────────────────────────────
TEAL_DARK    = "0D6B7E"
TEAL_MID     = "1A8FA0"
TEAL_LIGHT   = "B3E5EC"
GRAY_HEADER  = "455A64"
GRAY_DARK    = "37474F"
GREEN_DARK   = "1E5631"
GREEN_LIGHT  = "C6EFCE"
ORANGE_DARK  = "9C5700"
ORANGE_LIGHT = "FFEB9C"
PURPLE_DARK  = "4A148C"
PURPLE_LIGHT = "E1BEE7"
WHITE        = "FFFFFF"
GRAY_ROW     = "F2F2F2"

# ── Estilos compartidos ───────────────────────────────────────────────────────
_THIN_SIDE  = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE
)
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_GRAY  = PatternFill(start_color=GRAY_ROW, end_color=GRAY_ROW, fill_type="solid")


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")


def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_title(ws, row, text, hex_bg, font_color="FFFFFF", size=11, end_col=10):
    """Escribe una celda de titulo fusionada con fondo de color."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _font(bold=True, color=font_color, size=size)
    cell.fill = _fill(hex_bg)
    cell.alignment = _align(h="center", v="center")
    ws.row_dimensions[row].height = 28
    if end_col > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=end_col
        )


def write_header_row(ws, row, headers, hex_bg, font_color="FFFFFF", row_height=18):
    """Escribe una fila de encabezados con fondo de color."""
    ws.row_dimensions[row].height = row_height
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _font(bold=True, color=font_color, size=9)
        cell.fill = _fill(hex_bg)
        cell.alignment = _align(h="center", wrap=True)
        cell.border = THIN_BORDER


def write_data_row(ws, row, values, alt=False, number_cols=None, left_cols=None, formula_cols=None):
    """Escribe una fila de datos con alternancia de color y formatos especiales."""
    fill = FILL_GRAY if alt else FILL_WHITE
    number_cols  = number_cols  or []
    left_cols    = left_cols    or []
    formula_cols = formula_cols or {}  # dict: col -> formula string
    ws.row_dimensions[row].height = 15
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col)
        if col in formula_cols:
            cell.value = formula_cols[col]
        else:
            cell.value = v
        cell.fill = fill
        cell.border = THIN_BORDER
        if col in number_cols:
            cell.alignment = _align(h="right")
            cell.number_format = '#,##0'
        elif col in left_cols:
            cell.alignment = _align(h="left")
        else:
            cell.alignment = _align(h="center", wrap=True)


# ── Utilidades ────────────────────────────────────────────────────────────────
def sec_to_hms(seconds):
    """Convierte segundos a formato HH:MM:SS."""
    if not seconds:
        return "00:00:00"
    s = int(seconds)
    return f"{s // 3600:02d}:{(s % 3600) // 60:02d}:{s % 60:02d}"


def mes_nombre(n):
    """Devuelve el nombre del mes en espanol."""
    names = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    return names[n] if 1 <= n <= 12 else str(n)


def fecha_corta(iso):
    """Convierte ISO 8601 a formato dd/MM/yyyy."""
    if not iso:
        return ""
    parts = iso[:10].split("-")
    if len(parts) == 3:
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    return iso[:10]


def meds_texto(meds, excluir_categoria=None, solo_categoria=None):
    """
    Genera texto resumen de medicamentos para celdas del Resumen.
    Filtra por categoria si se indica.
    Formato: "Nombre (xN unidad), Nombre2 (xN2 unidad2)"
    """
    filtrados = []
    for m in meds:
        cat = m.get("categoria", "general")
        if excluir_categoria and cat == excluir_categoria:
            continue
        if solo_categoria and cat != solo_categoria:
            continue
        filtrados.append(f"{m['nombre']} (x{m['cantidad']} {m.get('unidad','u')})")
    return ", ".join(filtrados) if filtrados else ""


# ── Hoja 1: Resumen ───────────────────────────────────────────────────────────
def build_resumen(ws, data, mes, anio):
    """
    Una fila por sesion de cada paciente.
    Columnas: Ficha | Nombre | RUT | FECHA | HORAS DE SILLON |
              INSUMOS & MEDICAMENTOS | OTROS | TOTAL NETO | IVA | TOTAL
    """
    ws.sheet_view.showGridLines = False
    titulo = f"RESUMEN COP — {mes_nombre(mes).upper()} {anio}"
    write_title(ws, 1, titulo, TEAL_DARK, end_col=10)

    headers = [
        "Ficha", "Nombre", "RUT",
        "FECHA", "HORAS DE SILLON",
        "INSUMOS & MEDICAMENTOS", "OTROS",
        "TOTAL NETO", "IVA", "TOTAL"
    ]
    write_header_row(ws, 2, headers, TEAL_MID)

    row = 3
    for ficha_num, paciente in enumerate(data["pacientes"], 1):
        rut = paciente.get("rut") or paciente.get("pasaporte") or ""
        for sesion in paciente.get("sesiones", []):
            total_neto = sesion.get("totalSesion", 0)
            insumos    = meds_texto(sesion.get("medicamentos", []), excluir_categoria="otros")
            otros      = meds_texto(sesion.get("medicamentos", []), solo_categoria="otros")

            values = [
                ficha_num,
                paciente["nombreCompleto"],
                rut,
                fecha_corta(sesion.get("horaInicio", "")),
                sec_to_hms(sesion.get("duracionSegundos")),
                insumos,
                otros,
                total_neto,
                None,  # IVA — formula
                None,  # TOTAL — formula
            ]
            alt = (row % 2 == 0)
            # Formulas de IVA y TOTAL (cols 9 y 10)
            iva_formula   = f"=H{row}*0.19"
            total_formula = f"=H{row}+I{row}"
            write_data_row(
                ws, row, values,
                alt=alt,
                number_cols=[1, 8, 9, 10],
                left_cols=[2, 6, 7],
                formula_cols={9: iva_formula, 10: total_formula}
            )
            # Aplicar formato numerico a celdas de formula
            for fc in [9, 10]:
                ws.cell(row=row, column=fc).number_format = '#,##0'
                ws.cell(row=row, column=fc).fill = FILL_GRAY if alt else FILL_WHITE
                ws.cell(row=row, column=fc).border = THIN_BORDER
                ws.cell(row=row, column=fc).alignment = _align(h="right")
            row += 1

    if row == 3:
        ws.cell(row=3, column=1, value="Sin sesiones registradas en este periodo").font = _font(italic=True, color="999999")
        ws.merge_cells("A3:J3")
        row = 4

    # Fila de totales
    ws.row_dimensions[row].height = 18
    total_label = ws.cell(row=row, column=7, value="TOTALES")
    total_label.font = _font(bold=True, color=TEAL_DARK, size=10)
    total_label.fill = _fill(TEAL_LIGHT)
    total_label.alignment = _align(h="right")

    for col in [8, 9, 10]:
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}3:{col_letter}{row - 1})" if row > 3 else 0
        c = ws.cell(row=row, column=col, value=formula if row > 3 else 0)
        c.font = _font(bold=True, color=TEAL_DARK, size=10)
        c.fill = _fill(TEAL_LIGHT)
        c.number_format = '#,##0'
        c.alignment = _align(h="right")
        c.border = THIN_BORDER

    widths = [8, 30, 16, 12, 18, 40, 30, 14, 12, 14]
    for col, w in enumerate(widths, 1):
        set_col_width(ws, col, w)


# ── Hoja 2: Descripcion ───────────────────────────────────────────────────────
def build_descripcion(ws, data, mes, anio):
    """
    Catalogo de medicamentos del mes dividido en dos grupos:
    Cols A-C: DESCRIPCION | Codigo | Valor Unitario  (insumos normales)
    Cols D-E: separador vacio
    Cols F-H: OTROS | CODIGO | Honorarios           (categoria 'otros')
    """
    ws.sheet_view.showGridLines = False
    titulo = f"DESCRIPCION — {mes_nombre(mes).upper()} {anio}"
    write_title(ws, 1, titulo, GREEN_DARK, end_col=8)

    # Encabezados grupo izquierdo (A-C)
    for col, h in enumerate(["DESCRIPCION", "Codigo", "Valor Unitario"], 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = _font(bold=True, color="FFFFFF", size=9)
        cell.fill = _fill(GREEN_DARK)
        cell.alignment = _align(h="center", wrap=True)
        cell.border = THIN_BORDER
        ws.row_dimensions[2].height = 18

    # Encabezados grupo derecho (F-H)
    for col, h in enumerate(["OTROS", "CODIGO", "Honorarios"], 6):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = _font(bold=True, color="FFFFFF", size=9)
        cell.fill = _fill(PURPLE_DARK)
        cell.alignment = _align(h="center", wrap=True)
        cell.border = THIN_BORDER

    # Separar medicamentos: normales vs 'otros'
    normales = [m for m in data["medicamentos"] if m.get("categoria", "general") != "otros"]
    otros    = [m for m in data["medicamentos"] if m.get("categoria", "general") == "otros"]
    total_rows = max(len(normales), len(otros))

    for i in range(total_rows):
        row = 3 + i
        ws.row_dimensions[row].height = 15
        alt = (i % 2 == 0)

        # Grupo izquierdo
        if i < len(normales):
            med = normales[i]
            for col, v in enumerate([med["nombre"], med.get("codigoInterno") or "", med.get("costoTotal", 0) // max(med.get("cantidadTotal", 1), 1)], 1):
                c = ws.cell(row=row, column=col, value=v)
                c.fill = FILL_GRAY if alt else FILL_WHITE
                c.border = THIN_BORDER
                c.alignment = _align(h="left" if col == 1 else "center")
                if col == 3:
                    c.number_format = '#,##0'
                    c.alignment = _align(h="right")

        # Grupo derecho (sin cols D-E)
        if i < len(otros):
            med = otros[i]
            for col, v in enumerate([med["nombre"], med.get("codigoInterno") or "", med.get("costoTotal", 0) // max(med.get("cantidadTotal", 1), 1)], 6):
                c = ws.cell(row=row, column=col, value=v)
                c.fill = _fill(PURPLE_LIGHT) if alt else FILL_WHITE
                c.border = THIN_BORDER
                c.alignment = _align(h="left" if col == 6 else "center")
                if col == 8:
                    c.number_format = '#,##0'
                    c.alignment = _align(h="right")

    if total_rows == 0:
        ws.cell(row=3, column=1, value="Sin medicamentos en este periodo").font = _font(italic=True, color="999999")

    widths = [36, 16, 16, 4, 4, 36, 16, 16]
    for col, w in enumerate(widths, 1):
        set_col_width(ws, col, w)


# ── Hoja 3: Hora Sillon ───────────────────────────────────────────────────────
def build_hora_sillon(ws, data, mes, anio):
    """
    Catalogo de medicamentos con detalles tecnicos y de precio.
    Columnas: Descripcion | Cod Inte | Principio Activo | Laboratorio |
              Precio Neto | IVA | Recargo | Columna1
    """
    ws.sheet_view.showGridLines = False
    titulo = f"HORA SILLON — {mes_nombre(mes).upper()} {anio}"
    write_title(ws, 1, titulo, GRAY_DARK, end_col=8)

    headers = [
        "Descripcion", "Cod Inte", "Principio Activo",
        "Laboratorio", "Precio Neto", "IVA", "Recargo", "Columna1"
    ]
    write_header_row(ws, 2, headers, GRAY_HEADER)

    row = 3
    for i, med in enumerate(data["medicamentos"]):
        precio_neto = med.get("costoTotal", 0) // max(med.get("cantidadTotal", 1), 1)
        values = [
            med["nombre"],
            med.get("codigoInterno") or "",
            med.get("principioActivo") or "",
            med.get("laboratorio") or "",
            precio_neto,
            None,   # IVA — formula
            0,      # Recargo
            "",     # Columna1
        ]
        alt = (i % 2 == 0)
        iva_formula = f"=E{row}*0.19"
        write_data_row(
            ws, row, values,
            alt=alt,
            number_cols=[5, 6, 7],
            left_cols=[1, 3, 4],
            formula_cols={6: iva_formula}
        )
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=6).fill   = FILL_GRAY if alt else FILL_WHITE
        ws.cell(row=row, column=6).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = _align(h="right")
        row += 1

    if not data["medicamentos"]:
        ws.cell(row=3, column=1, value="Sin medicamentos en este periodo").font = _font(italic=True, color="999999")

    widths = [36, 14, 24, 22, 14, 12, 12, 14]
    for col, w in enumerate(widths, 1):
        set_col_width(ws, col, w)


# ── Hoja 4: Prevision ─────────────────────────────────────────────────────────
def build_prevision(ws, data, mes, anio):
    """
    Lista de pacientes del mes con su prevision.
    La columna Prevision tiene un dropdown con todas las isapres/fonasa disponibles.
    """
    ws.sheet_view.showGridLines = False
    titulo = f"PREVISION — {mes_nombre(mes).upper()} {anio}"
    write_title(ws, 1, titulo, ORANGE_DARK, end_col=4)

    headers = ["Ficha", "Nombre", "RUT", "Prevision"]
    write_header_row(ws, 2, headers, ORANGE_DARK)

    # Dropdown de previsiones en columna D (col 4)
    prev_formula = '"' + ",".join(PREVISIONES) + '"'
    dv = DataValidation(
        type="list",
        formula1=prev_formula,
        allow_blank=True,
        showDropDown=False,   # False = mostrar flecha del desplegable
    )
    ws.add_data_validation(dv)
    dv.sqref = f"D3:D1000"

    row = 3
    for ficha_num, paciente in enumerate(data["pacientes"], 1):
        rut  = paciente.get("rut") or paciente.get("pasaporte") or ""
        prev = paciente.get("prevision") or ""
        alt  = (ficha_num % 2 == 0)
        write_data_row(ws, row, [ficha_num, paciente["nombreCompleto"], rut, prev],
                       alt=alt, number_cols=[1], left_cols=[2])
        row += 1

    if not data["pacientes"]:
        ws.cell(row=3, column=1, value="Sin pacientes en este periodo").font = _font(italic=True, color="999999")
        ws.merge_cells("A3:D3")

    widths = [8, 34, 16, 24]
    for col, w in enumerate(widths, 1):
        set_col_width(ws, col, w)


# ── Hojas 5-40: Hoja1 ... Hoja36 (vacias, datos a definir) ───────────────────
def build_hoja_vacia(ws, num):
    """Hoja en blanco numerada. El contenido se definira posteriormente."""
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value=f"Hoja {num}").font = _font(bold=True, color="AAAAAA", size=9)
    ws.cell(row=2, column=1, value="Contenido pendiente de definicion.").font = _font(italic=True, color="CCCCCC", size=8)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 3:
        print("Uso: python3 generate_cop.py <input.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    input_file  = sys.argv[1]
    output_file = sys.argv[2]

    with open(input_file, "r", encoding="utf-8") as f:
        payload = json.load(f)

    data = payload["data"]
    mes  = int(payload["mes"])
    anio = int(payload["anio"])

    wb = Workbook()
    wb.remove(wb.active)  # eliminar la hoja por defecto

    # ── 4 hojas fijas ──
    build_resumen(    wb.create_sheet("Resumen"),      data, mes, anio)
    build_descripcion(wb.create_sheet("Descripcion"),  data, mes, anio)
    build_hora_sillon(wb.create_sheet("Hora Sillon"),  data, mes, anio)
    build_prevision(  wb.create_sheet("Prevision"),    data, mes, anio)

    # ── 36 hojas en blanco (Hoja1 - Hoja36) ──
    for i in range(1, 37):
        build_hoja_vacia(wb.create_sheet(f"Hoja{i}"), i)

    wb.save(output_file)
    print(f"OK:{output_file}", flush=True)


if __name__ == "__main__":
    main()
